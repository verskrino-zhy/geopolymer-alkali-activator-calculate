# -*- coding: utf-8 -*-
"""
Geopolymer Alkali Activator Calculator (UI Polished + Footer)
地聚物碱激发剂参数计算软件 · 视觉优化版（单文件）

依赖：
  pip install PySide6 openpyxl pandas

本版优化：
- 去除“保存图标”按钮。
- 模块标题采用高光渐变徽章，字号更大、层次更清晰。
- 卡片与按钮加入“浮起”阴影（elevation）与圆角，减少扁平感。
- 主次按钮采用不同渐变；悬停/按压有微互动反馈。
- 输入变更自动联动计算（仍保留“计算”按钮）。
- 右下角新增署名与链接：朱桓毅制作，仅供参考，欢迎讨论。www.verskrino.com

核心计算与变量含义与前版一致。
"""
from __future__ import annotations

import sys
import math
from dataclasses import dataclass
from typing import Dict, Tuple

from PySide6.QtCore import Qt
from PySide6.QtGui import (
    QAction, QColor, QDoubleValidator, QFont, QIcon, QLinearGradient,
    QPainter, QPainterPath, QPen, QPixmap, QBrush
)
from PySide6.QtWidgets import (
    QApplication, QWidget, QVBoxLayout, QHBoxLayout, QFormLayout, QGroupBox,
    QLineEdit, QLabel, QPushButton, QFileDialog, QMessageBox, QFrame,
    QGraphicsDropShadowEffect
)

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# ---------------------- 常量与配色 ----------------------
K62_60 = 62.0/60.0           # = 31/30
NAOH_TO_NA2O = 62.0/80.0     # 0.775

UI_COLORS = {
    'bg': '#f6f7fb',
    'text': '#0f172a',
    'muted': '#6b7280',
    'primary1': '#38bdf8',  # 蓝 渐变起
    'primary2': '#0ea5e9',  # 蓝 渐变止
    'target1':  '#c084fc',  # 紫 渐变起
    'target2':  '#a855f7',  # 紫 渐变止
    'process1': '#fbbf24',  # 橙 渐变起
    'process2': '#f59e0b',  # 橙 渐变止
    'key1':     '#34d399',  # 绿 渐变起
    'key2':     '#22c55e',  # 绿 渐变止
    'card':    '#ffffff',
    'border':  '#e5e7eb'
}

# ---------------------- 工具函数 ----------------------

def parse_percent(value_str: str) -> float:
    """将输入解析为比例（0~1）。支持：0~1 或 0~100。"""
    s = (value_str or "").strip()
    if not s:
        return float("nan")
    try:
        v = float(s)
    except ValueError:
        return float("nan")
    if v > 1.0:
        v = v / 100.0
    return v


def parse_float(value_str: str) -> float:
    s = (value_str or "").strip()
    if not s:
        return float("nan")
    try:
        return float(s)
    except ValueError:
        return float("nan")


# ---------------------- 数据结构 ----------------------
@dataclass
class Inputs:
    A: float  # 原料体系总质量（g）
    C: float  # 水玻璃中二氧化硅百分比（比例）
    D: float  # 水玻璃中氧化钠百分比（比例）
    O: float  # 新体系的碱激发剂模数（目标）
    Q: float  # 最终碱当量（目标）
    R: float  # 固液比（目标）


@dataclass
class Results:
    # 关键结果
    B: float  # 需添加水玻璃的总质量（g）
    E: float  # 需要添加的氢氧化钠质量（g）
    F: float  # 需要添加的水量（g）

    # 过程结果
    G: float  # 新体系中二氧化硅百分比（比例）
    H: float  # 新体系中氧化钠百分比（比例）
    I: float  # 新体系液体密度（g/cm^3）
    J: float  # 新体系液体质量（g）
    K: float  # 水玻璃模数（验证用）
    L: float  # 水玻璃中二氧化硅质量（g）
    M: float  # 水玻璃中氧化钠质量（g）
    N: float  # 添加的氧化钠质量换算量（g）
    O: float  # 新体系的碱激发剂模数（回算）
    P: float  # 初始碱当量
    Q: float  # 最终碱当量（回算）
    R: float  # 固液比（回算）


# ---------------------- 主界面 ----------------------
class GeoActivatorApp(QWidget):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("地聚物碱激发剂参数计算软件")
        self.setMinimumWidth(1000)
        self._app_icon = self._make_app_icon(256)
        self.setWindowIcon(QIcon(self._app_icon))
        self._build_ui()
        self._apply_styles()

    # ---------- 布局 ----------
    def _build_ui(self):
        main = QVBoxLayout(self)
        main.setContentsMargins(16, 16, 16, 8)
        main.setSpacing(12)

        # 顶部栏
        header = QHBoxLayout()
        logo_lbl = QLabel()
        logo_lbl.setPixmap(self._app_icon.scaled(48, 48, Qt.KeepAspectRatio, Qt.SmoothTransformation))
        title_box = QVBoxLayout()
        h1 = QLabel("地聚物碱激发剂参数计算软件"); h1.setObjectName("appTitle")
        h2 = QLabel("输入基础与目标参数，自动反算关键配比与过程指标"); h2.setObjectName("appSubtitle")
        title_box.addWidget(h1); title_box.addWidget(h2)
        header.addWidget(logo_lbl)
        header.addLayout(title_box)
        header.addStretch(1)

        # 操作按钮
        self.btn_export = QPushButton("导出为 Excel…"); self.btn_export.setObjectName("accentBtn"); self._elevate(self.btn_export)
        self.btn_calc   = QPushButton("计算");         self.btn_calc.setObjectName("primaryBtn"); self._elevate(self.btn_calc)
        header.addWidget(self.btn_export)
        header.addWidget(self.btn_calc)
        main.addLayout(header)

        # 输入区
        io_row = QHBoxLayout(); main.addLayout(io_row)
        self.gb_base = QGroupBox("基础参数");  self.gb_base.setObjectName("gb_base")
        self.gb_targets = QGroupBox("目标参数"); self.gb_targets.setObjectName("gb_targets")
        io_row.addWidget(self._wrap_card(self.gb_base, tint=(UI_COLORS['primary1'], UI_COLORS['primary2'])), 1)
        io_row.addWidget(self._wrap_card(self.gb_targets, tint=(UI_COLORS['target1'], UI_COLORS['target2'])), 1)

        # 基础参数表单
        form_base = QFormLayout(); form_base.setLabelAlignment(Qt.AlignRight)
        self.le_A = QLineEdit(); self._as_number(self.le_A)
        self.le_C = QLineEdit(); self._as_number(self.le_C)
        self.le_D = QLineEdit(); self._as_number(self.le_D)
        form_base.addRow(self._label_with_note("原料体系总质量（g）", "指地聚物固体原料总质量，如粉煤灰、煤矸石等。"), self.le_A)
        form_base.addRow(self._label_with_note("水玻璃中二氧化硅百分比（%）", "输入可为 0~1 或 0~100；>1 自动除以 100。"), self.le_C)
        form_base.addRow(self._label_with_note("水玻璃中氧化钠百分比（%）", "输入可为 0~1 或 0~100；>1 自动除以 100。"), self.le_D)
        self.gb_base.setLayout(form_base)

        # 目标参数表单
        form_tar = QFormLayout(); form_tar.setLabelAlignment(Qt.AlignRight)
        self.le_O = QLineEdit(); self._as_number(self.le_O)
        self.le_Q = QLineEdit(); self._as_number(self.le_Q)
        self.le_R = QLineEdit(); self._as_number(self.le_R)
        form_tar.addRow(self._label_with_note("新体系的碱激发剂模数", "最终体系 n(Na2O)/n(SiO2)。"), self.le_O)
        form_tar.addRow(self._label_with_note("最终碱当量", "m(Na2O@水玻璃)+m(Na2O@NaOH换算) 与固体质量之比；液体不在分母。"), self.le_Q)
        form_tar.addRow(self._label_with_note("固液比 (S/L)", "固体原料总质量 / 所有液体总质量。"), self.le_R)
        self.gb_targets.setLayout(form_tar)

        # 信息条
        self.msg = QLabel(); self.msg.setObjectName("message")
        main.addWidget(self.msg)

        # 输出区
        out_row = QHBoxLayout(); main.addLayout(out_row)
        self.gb_proc = QGroupBox("过程结果"); self.gb_proc.setObjectName("gb_process")
        self.gb_key  = QGroupBox("关键结果"); self.gb_key.setObjectName("gb_key")
        out_row.addWidget(self._wrap_card(self.gb_proc, tint=(UI_COLORS['process1'], UI_COLORS['process2'])), 1)
        out_row.addWidget(self._wrap_card(self.gb_key,  tint=(UI_COLORS['key1'], UI_COLORS['key2'])), 1)

        self.proc_labels: Dict[str, QLineEdit] = {}
        self.key_labels: Dict[str, QLineEdit] = {}

        form_proc = QFormLayout(); form_proc.setLabelAlignment(Qt.AlignRight)
        self._add_readonly(form_proc, self.proc_labels, "新体系中二氧化硅百分比（%）", key="G")
        self._add_readonly(form_proc, self.proc_labels, "新体系中氧化钠百分比（%）", key="H")
        self._add_readonly(form_proc, self.proc_labels, "新体系液体密度（g/cm³）", key="I")
        self._add_readonly(form_proc, self.proc_labels, "新体系液体质量（g）", key="J")
        self._add_readonly(form_proc, self.proc_labels, "水玻璃模数（验证用）", key="K")
        self._add_readonly(form_proc, self.proc_labels, "水玻璃中二氧化硅质量（g）", key="L")
        self._add_readonly(form_proc, self.proc_labels, "水玻璃中氧化钠质量（g）", key="M")
        self._add_readonly(form_proc, self.proc_labels, "添加的氧化钠质量换算量（g）", key="N")
        self._add_readonly(form_proc, self.proc_labels, "初始碱当量", key="P")
        self.gb_proc.setLayout(form_proc)

        form_key = QFormLayout(); form_key.setLabelAlignment(Qt.AlignRight)
        self._add_readonly(form_key, self.key_labels, "需添加水玻璃的总质量（g）", key="B")
        self._add_readonly(form_key, self.key_labels, "需要添加的氢氧化钠质量（g）", key="E")
        self._add_readonly(form_key, self.key_labels, "需要添加的水量（g）", key="F")
        self._add_readonly(form_key, self.key_labels, "新体系的碱激发剂模数(回算)", key="O")
        self._add_readonly(form_key, self.key_labels, "最终碱当量(回算)", key="Q")
        self._add_readonly(form_key, self.key_labels, "固液比(回算)", key="R")
        self.gb_key.setLayout(form_key)

        # 备注
        note = QLabel(
            "<b>备注说明：</b><br>"
            "1) 原料体系总质量：指地聚物固体原料总质量。<br>"
            "2) 新体系液体密度：水玻璃+NaOH 后整体液体密度。<br>"
            "3) 水玻璃模数（验证用）：由输入成分回算，用于核查。<br>"
            "4) 最终碱当量：液体不计入分母。<br>"
            "5) 新体系碱激发剂模数：最终体系 n(Na2O)/n(SiO2)。<br>"
            "6) 固液比：固体/所有液体，总液体=水玻璃+NaOH+外加水。")
        note.setObjectName("notes")
        main.addWidget(note)

        # 右下角署名 + 链接
        footer = QLabel('朱桓毅制作，仅供参考，欢迎讨论。<a href="https://www.verskrino.com">www.verskrino.com</a>')
        footer.setObjectName("footer")
        footer.setTextFormat(Qt.RichText)
        footer.setOpenExternalLinks(True)
        footer.setAlignment(Qt.AlignRight)
        main.addWidget(footer)

        # 事件绑定
        self.btn_calc.clicked.connect(self.on_calc)
        self.btn_export.clicked.connect(self.on_export)
        for le in [self.le_A, self.le_C, self.le_D, self.le_O, self.le_Q, self.le_R]:
            le.textChanged.connect(self._auto_calc)

        # 右键菜单：示例
        act_demo = QAction("填充示例", self)
        act_demo.triggered.connect(self.fill_demo)
        self.addAction(act_demo)
        self.setContextMenuPolicy(Qt.ActionsContextMenu)

    # ---------- 卡片包装 + 阴影 ----------
    def _wrap_card(self, inner: QGroupBox, tint: tuple[str, str]) -> QWidget:
        w = QWidget()
        lay = QVBoxLayout(w); lay.setContentsMargins(0, 0, 0, 0); lay.addWidget(inner)
        # 卡片阴影
        self._elevate(w, blur=28, offset_y=2, alpha=60)
        # 将标题徽章的渐变写入对象名，供样式表区分
        inner.setProperty('badgeStart', tint[0]); inner.setProperty('badgeEnd', tint[1])
        return w

    def _elevate(self, widget: QWidget, blur: int = 22, offset_y: int = 3, alpha: int = 50):
        eff = QGraphicsDropShadowEffect(widget)
        eff.setBlurRadius(blur)
        eff.setXOffset(0); eff.setYOffset(offset_y)
        eff.setColor(QColor(0, 0, 0, alpha))
        widget.setGraphicsEffect(eff)

    def _as_number(self, le: QLineEdit):
        le.setPlaceholderText("输入数值…")
        le.setValidator(QDoubleValidator(bottom=-1e12, top=1e12, decimals=6))
        le.setClearButtonEnabled(True)
        le.setFixedHeight(36)

    def _label_with_note(self, text: str, note: str) -> QLabel:
        lab = QLabel(text)
        lab.setToolTip(note)
        return lab

    def _add_readonly(self, form: QFormLayout, store: Dict[str, QLineEdit], label: str, key: str):
        le = QLineEdit(); le.setReadOnly(True); le.setObjectName("ro"); le.setFixedHeight(34)
        form.addRow(QLabel(label), le)
        store[key] = le

    # ---------- 样式表（标题徽章 + 渐变按钮） ----------
    def _apply_styles(self):
        # 使用 qlineargradient 做徽章和按钮渐变；卡片使用圆角+边线；
        qss = f"""
        QWidget {{ background: {UI_COLORS['bg']}; color: {UI_COLORS['text']}; font-size: 14px; }}
        #appTitle {{ font-size: 24px; font-weight: 800; margin-left: 8px; }}
        #appSubtitle {{ color: {UI_COLORS['muted']}; margin-left: 8px; }}

        /* 卡片与分组 */
        QGroupBox {{
            background: {UI_COLORS['card']};
            border: 1px solid {UI_COLORS['border']};
            border-radius: 14px; margin-top: 36px; padding: 14px 14px 14px 14px;
        }}
        QGroupBox::title {{
            subcontrol-origin: margin; subcontrol-position: top left;
            padding: 8px 14px; border-radius: 10px; color: white;
            font-size: 18px; font-weight: 800; letter-spacing: 0.3px;
        }}
        /* 为不同分组标题植入横向渐变 */
        QGroupBox#gb_base::title {{
            background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                stop:0 {UI_COLORS['primary1']}, stop:1 {UI_COLORS['primary2']});
        }}
        QGroupBox#gb_targets::title {{
            background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                stop:0 {UI_COLORS['target1']}, stop:1 {UI_COLORS['target2']});
        }}
        QGroupBox#gb_process::title {{
            background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                stop:0 {UI_COLORS['process1']}, stop:1 {UI_COLORS['process2']});
        }}
        QGroupBox#gb_key::title {{
            background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                stop:0 {UI_COLORS['key1']}, stop:1 {UI_COLORS['key2']});
        }}

        /* 输入控件 */
        QLineEdit {{ background: white; border: 1px solid {UI_COLORS['border']}; border-radius: 10px; padding: 6px 10px; }}
        QLineEdit:focus {{ border: 1px solid {UI_COLORS['primary2']}; }}
        QLineEdit#ro {{ background: #f9fafb; color: #111827; }}

        QLabel#message {{ color: #c62828; min-height: 22px; margin: 4px 2px; }}
        QLabel#notes {{ color: {UI_COLORS['muted']}; line-height: 1.55; margin-top: 8px; }}
        QLabel#footer {{ color: {UI_COLORS['muted']}; font-size: 12px; margin: 8px 2px 0 0; }}

        /* 按钮 —— 圆角、渐变、阴影（阴影由 QGraphicsDropShadowEffect 提供） */
        QPushButton {{ border: none; border-radius: 12px; padding: 10px 18px; font-weight: 700; font-size: 15px; }}
        QPushButton#primaryBtn {{
            color: white;
            background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                stop:0 {UI_COLORS['primary1']}, stop:1 {UI_COLORS['primary2']});
        }}
        QPushButton#primaryBtn:hover {{ filter: brightness(1.05); }}
        QPushButton#primaryBtn:pressed {{ padding-top: 12px; padding-bottom: 8px; }}

        QPushButton#accentBtn {{
            color: #111827;
            background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                stop:0 {UI_COLORS['process1']}, stop:1 {UI_COLORS['process2']});
        }}
        QPushButton#accentBtn:hover {{ filter: brightness(1.04); }}
        QPushButton#accentBtn:pressed {{ padding-top: 12px; padding-bottom: 8px; }}
        """
        self.setStyleSheet(qss)

    # ---------- 计算逻辑 ----------
    def _collect_inputs(self) -> Tuple[Inputs | None, str]:
        A = parse_float(self.le_A.text())
        C = parse_percent(self.le_C.text())
        D = parse_percent(self.le_D.text())
        O = parse_float(self.le_O.text())
        Qv = parse_float(self.le_Q.text())
        R = parse_float(self.le_R.text())

        err = []
        if any(math.isnan(x) for x in [A, C, D, O, Qv, R]):
            err.append("存在空值或非数字输入。")
        if A <= 0: err.append("原料体系总质量 A 必须 > 0。")
        if not (0 < C < 1): err.append("水玻璃二氧化硅百分比 C 需在 (0,1) 内（可输入 0~100，将自动换算）。")
        if not (0 < D < 1): err.append("水玻璃氧化钠百分比 D 需在 (0,1) 内（可输入 0~100，将自动换算）。")
        if O <= 0: err.append("目标模数 O 必须 > 0。")
        if Qv <= 0: err.append("最终碱当量 Q 必须 > 0。")
        if R <= 0: err.append("固液比 R 必须 > 0。")
        if C + D >= 1: err.append("水玻璃中 SiO2 与 Na2O 百分比之和需小于 100%。")
        if err:
            return None, "；".join(err)
        return Inputs(A, C, D, O, Qv, R), ""

    def _compute(self, inp: Inputs) -> Tuple[Results | None, str]:
        A, C, D, O, Qv, R = inp.A, inp.C, inp.D, inp.O, inp.Q, inp.R
        if C <= 0: return None, "C=0，无法计算 B。"
        B = (O * Qv * A) / (C * K62_60)
        E = (Qv * A - B * D) / NAOH_TO_NA2O
        F = A / R - (B + E)
        if any(x != x for x in [B, E, F]): return None, "出现 NaN 结果，请检查输入。"
        if B <= 0: return None, "反算得到 B≤0，请调整目标参数（可能 O、Q 过小或 C 过大）。"
        if E < 0: return None, "反算得到 E<0，请调整目标参数（可能 Q 偏小或 D 偏大）。"
        if F < 0: return None, "反算得到 F<0，请调整目标参数或固液比 R。"

        L = B * C; M = B * D; N = E * NAOH_TO_NA2O
        J = B + E
        G = L / (B + E) if (B + E) > 0 else float("nan")
        H = (M + N) / (B + E) if (B + E) > 0 else float("nan")
        I = 1.0 / (((1.0 - G - H)/0.998) + (G/2.2) + (H/2.27)) if not any(math.isnan(x) for x in [G, H]) else float("nan")
        K = (L/60.0) / (M/62.0) if M > 0 else float("nan")
        O_back = (L/60.0) / ((M/62.0) + (N/62.0)) if (M + N) > 0 else float("nan")
        P = M / A
        Q_back = (M + N) / A
        R_back = A / (J + F) if (J + F) > 0 else float("nan")
        return Results(B, E, F, G, H, I, J, K, L, M, N, O_back, P, Q_back, R_back), ""

    # ---------- 事件 ----------
    def on_calc(self):
        self.msg.clear()
        inputs, err = self._collect_inputs()
        if err:
            self.msg.setText(err); return
        res, err = self._compute(inputs)
        if err:
            self.msg.setText(err); return
        self._fill_outputs(res)

    def _auto_calc(self):
        if any(le.text().strip()=='' for le in [self.le_A, self.le_C, self.le_D, self.le_O, self.le_Q, self.le_R]):
            return
        self.on_calc()

    def _fill_outputs(self, r: Results):
        # 过程
        self.proc_labels["G"].setText(f"{r.G*100:.3f}")
        self.proc_labels["H"].setText(f"{r.H*100:.3f}")
        self.proc_labels["I"].setText(f"{r.I:.4f}")
        self.proc_labels["J"].setText(f"{r.J:.3f}")
        self.proc_labels["K"].setText(f"{r.K:.4f}")
        self.proc_labels["L"].setText(f"{r.L:.3f}")
        self.proc_labels["M"].setText(f"{r.M:.3f}")
        self.proc_labels["N"].setText(f"{r.N:.3f}")
        self.proc_labels["P"].setText(f"{r.P:.5f}")
        # 关键
        self.key_labels["B"].setText(f"{r.B:.3f}")
        self.key_labels["E"].setText(f"{r.E:.3f}")
        self.key_labels["F"].setText(f"{r.F:.3f}")
        self.key_labels["O"].setText(f"{r.O:.5f}")
        self.key_labels["Q"].setText(f"{r.Q:.5f}")
        self.key_labels["R"].setText(f"{r.R:.5f}")

    # ---------- 导出 Excel ----------
    def on_export(self):
        inputs, err = self._collect_inputs()
        if err:
            QMessageBox.warning(self, "输入有误", err); return
        res, err = self._compute(inputs)
        if err:
            QMessageBox.warning(self, "无法导出", err); return
        path, _ = QFileDialog.getSaveFileName(self, "导出为 Excel", "计算结果.xlsx", "Excel 文件 (*.xlsx)")
        if not path: return
        try:
            self._export_excel(inputs, res, path)
        except Exception as e:
            QMessageBox.critical(self, "导出失败", f"{e}"); return
        QMessageBox.information(self, "完成", "已成功导出 Excel。")

    def _export_excel(self, inp: Inputs, r: Results, path: str):
        wb = Workbook(); ws = wb.active; ws.title = "Sheet1"
        headers = [
            "原料体系总质量（g）",
            "需添加水玻璃的总质量（g）",
            "水玻璃中二氧化硅百分比（%）",
            "水玻璃中氧化钠百分比（%）",
            "需要添加的氢氧化钠质量（g）",
            "需要添加的水量（g）",
            "新体系中二氧化硅百分比（%）",
            "新体系中氧化钠百分比（%）",
            "新体系液体密度（g/立方厘米）",
            "新体系液体质量（g）",
            "水玻璃模数（验证用）",
            "水玻璃中二氧化硅质量（g）",
            "水玻璃中氧化钠质量（g）",
            "添加的氧化钠质量换算量（g）",
            "新体系的碱激发剂模数",
            "初始碱当量",
            "最终碱当量",
            "固液比",
        ]
        varnames = [
            "m原", "m总", "Wsio2", "Wna2o", "mnaoh", "mh", "W新sio2", "W新na2o",
            "ρ新", "m液新", "M水玻璃", "m_SiO2", "m_Na2O@玻璃", "m_Na2O@NaOH",
            "M新", "N初", "N终", "S/L"
        ]
        ws.append(headers)
        ws.append(varnames)
        row3 = [inp.A, r.B, inp.C, inp.D, r.E, r.F, r.G, r.H, r.I, r.J, r.K, r.L, r.M, r.N, r.O, r.P, r.Q, r.R]
        ws.append([float(x) if isinstance(x, (int, float)) else x for x in row3])
        for col in range(1, len(headers)+1):
            ws.cell(row=1, column=col).font = ws.cell(row=1, column=col).font.copy(bold=True)
            ws.column_dimensions[get_column_letter(col)].width = 18
        wb.save(path)

    # ---------- 应用图标（程序内绘制） ----------
    def _make_app_icon(self, size: int = 256) -> QPixmap:
        pm = QPixmap(size, size); pm.fill(Qt.transparent)
        p = QPainter(pm); p.setRenderHints(QPainter.Antialiasing | QPainter.TextAntialiasing)
        # 背景圆渐变
        grad = QLinearGradient(0, 0, size, size)
        grad.setColorAt(0.0, QColor('#1d4ed8'))
        grad.setColorAt(1.0, QColor('#0ea5e9'))
        p.setBrush(QBrush(grad)); p.setPen(Qt.NoPen)
        p.drawEllipse(0, 0, size, size)
        # 烧杯
        path = QPainterPath(); inset = size*0.18; top = size*0.22; bottom = size*0.78; neck = size*0.40
        path.moveTo(inset, top); path.lineTo(size-inset, top); path.lineTo(size-inset-8, neck)
        path.lineTo(size-inset-18, bottom); path.lineTo(inset+18, bottom); path.lineTo(inset+8, neck); path.closeSubpath()
        p.setBrush(QColor(255,255,255,210)); p.setPen(QPen(QColor(255,255,255,240), 3)); p.drawPath(path)
        # 液面
        fluid = QPainterPath(); y_mid = size*0.58
        fluid.moveTo(inset+20, y_mid+10); fluid.lineTo(size-inset-20, y_mid-6)
        fluid.lineTo(size-inset-26, bottom-8); fluid.lineTo(inset+26, bottom-8); fluid.closeSubpath()
        p.setBrush(QColor('#22d3ee')); p.setPen(QPen(QColor('#0891b2'), 2)); p.drawPath(fluid)
        # 公式文本（Na2O / SiO2）
        def draw_formula(x, y, base, sub, tail, color):
            f = QFont('Segoe UI', int(size*0.13), QFont.DemiBold); p.setFont(f); p.setPen(QColor(color))
            p.drawText(x, y, base)
            f2 = QFont('Segoe UI', int(size*0.08)); p.setFont(f2); p.drawText(x + int(size*0.10), y + int(size*0.03), sub)
            f3 = QFont('Segoe UI', int(size*0.13), QFont.DemiBold); p.setFont(f3); p.drawText(x + int(size*0.14), y, tail)
        draw_formula(int(size*0.28), int(size*0.40), 'Na', '2', 'O', '#0b132a')
        draw_formula(int(size*0.34), int(size*0.66), 'Si', '2', 'O', '#0b132a')
        p.end(); return pm

    # ---------- 示例 ----------
    def fill_demo(self):
        self.le_A.setText("200")
        self.le_C.setText("30")
        self.le_D.setText("13.5")
        self.le_O.setText("1.5")
        self.le_Q.setText("0.15")
        self.le_R.setText("1.5")
        self.on_calc()


# ---------------------- 程序入口 ----------------------
if __name__ == "__main__":
    app = QApplication(sys.argv)
    w = GeoActivatorApp()
    w.show()
    sys.exit(app.exec())
